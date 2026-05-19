import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── colour palette ──────────────────────────────────────────────────────────
C = {
    "header_bg":   "1A1A2E", "header_fg": "FFFFFF",
    "m1": "16213E", "m2": "0F3460", "m3": "533483",
    "m4": "E94560", "m5": "0F3460", "m6": "16213E",
    "cat_bg": "E8F4FD", "alt_bg": "F8FFFE",
    "border": "CCCCCC",
    "p1": "FF6B6B", "p2": "FFA500", "p3": "4CAF50",
    "ph_bg": "FFF9C4",
    "done": "C8E6C9", "wip": "FFF9C4", "todo": "FFCCBC",
}

def fill(hex_):  return PatternFill("solid", fgColor=hex_)
def font(hex_="000000", bold=False, sz=11):
    return Font(color=hex_, bold=bold, size=sz, name="Calibri")
def border():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)
def align(h="left", v="center", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def hdr(ws, row, col, val, bg, fg="FFFFFF", bold=True, sz=11, h="center"):
    c = ws.cell(row, col, val)
    c.fill = fill(bg); c.font = font(fg, bold, sz)
    c.alignment = align(h); c.border = border()
    return c

def cell(ws, row, col, val, bg=None, fg="000000", bold=False, h="left"):
    c = ws.cell(row, col, val)
    if bg: c.fill = fill(bg)
    c.font = font(fg, bold)
    c.alignment = align(h); c.border = border()
    return c

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 – MASTER ROADMAP
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "📅 Master Roadmap"

# title
ws1.merge_cells("A1:M1")
t = ws1["A1"]
t.value = "🏥  VitalHealth – Digital Twin  |  6-Month Development Roadmap  (May 2026 – Oct 2026)"
t.fill = fill(C["header_bg"]); t.font = font(C["header_fg"], True, 16)
t.alignment = align("center"); ws1.row_dimensions[1].height = 36

# month labels
months = ["May 2026","Jun 2026","Jul 2026","Aug 2026","Sep 2026","Oct 2026"]
month_cols = [C["m1"],C["m2"],C["m3"],C["m4"],C["m5"],C["m6"]]
for i,(m,mc) in enumerate(zip(months,month_cols)):
    ws1.merge_cells(start_row=2, start_column=7+i, end_row=2, end_column=7+i)
    hdr(ws1, 2, 7+i, m, mc, sz=10)

# column headers
cols = ["#","Category","Feature / Task","Reason","Priority","Effort (days)",
        "May","Jun","Jul","Aug","Sep","Oct","Status"]
col_w = [4,20,42,45,10,12,8,8,8,8,8,8,12]
for i,(col,w) in enumerate(zip(cols,col_w),1):
    hdr(ws1, 3, i, col, "2C3E50")
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.row_dimensions[2].height = 22
ws1.row_dimensions[3].height = 20
ws1.freeze_panes = "A4"

prio_bg = {"🔴 Critical":C["p1"],"🟠 High":C["p2"],"🟢 Medium":C["p3"]}

rows = [
    # (cat, task, reason, prio, effort_days, gantt_months[0..5], status)
    # 1 = filled, 0 = empty
    # ── PHASE 1: Stability & Polish (May) ───────────────────────────────────
    ("PHASE 1: Stability & Core Polish","","","","","","","","","","","",""),
    ("BioGears Engine","Fix silent crash & pathing bugs","Prevents application from crashing unexpectedly during simulation","🔴 Critical",10,[1,0,0,0,0,0],"✅ Done"),
    ("BioGears Engine","Chunked time-advancement (no hangs)","Ensures the UI remains responsive during long simulation calculations","🔴 Critical",8,[1,0,0,0,0,0],"✅ Done"),
    ("BioGears Engine","XML schema validation pipeline","Validates engine configurations before runtime to prevent bad states","🔴 Critical",6,[1,0,0,0,0,0],"✅ Done"),
    ("BioGears Engine","Concurrency guard (no overlapping jobs)","Avoids engine race conditions when multiple actions are triggered","🟠 High",5,[1,0,0,0,0,0],"✅ Done"),
    ("BioGears Engine","State backup & rollback on failure","Ensures user simulation progress is not lost if a crash occurs","🟠 High",5,[1,0,0,0,0,0],"✅ Done"),
    ("Mobile App","Bold AI response rendering","Improves readability of AI-generated insights for users","🟢 Medium",2,[1,0,0,0,0,0],"✅ Done"),
    ("Mobile App","Background job persistence on restart","Keeps simulation tasks running even if the user closes the app","🟠 High",4,[1,1,0,0,0,0],"🔄 In Progress"),

    # ── PHASE 2: Simulation Accuracy (Jun) ──────────────────────────────────
    ("PHASE 2: Simulation Accuracy","","","","","","","","","","","",""),
    ("BioGears Engine","Multi-substance concurrent dosing support","Allows realistic scenarios where a patient takes multiple medications","🔴 Critical",10,[0,1,0,0,0,0],"📋 Planned"),
    ("BioGears Engine","Full substance registry (50+ drugs)","Expands the medical accuracy and utility of the digital twin","🟠 High",8,[0,1,0,0,0,0],"📋 Planned"),
    ("BioGears Engine","Pharmacokinetic unit validation","Ensures drug absorption and clearance match real-world medical data","🔴 Critical",6,[0,1,0,0,0,0],"📋 Planned"),
    ("BioGears Engine","Real-time vitals streaming (SSE)","Provides instant, continuous physiological feedback to the user","🟠 High",7,[0,1,1,0,0,0],"📋 Planned"),
    ("BioGears Engine","Blood glucose & insulin model","Crucial for diabetic users to predict sugar levels based on diet","🟠 High",10,[0,1,1,0,0,0],"📋 Planned"),
    ("AI / Analytics","Physiology anomaly detection (ML)","Identifies abnormal health patterns before they become critical","🟠 High",12,[0,1,1,0,0,0],"📋 Planned"),

    # ── PHASE 3: Wearable & Data Integration (Jul) ──────────────────────────
    ("PHASE 3: Wearable & Data Integration","","","","","","","","","","","",""),
    ("Wearables","Apple Health / Google Fit full sync","Automatically populates the twin with steps, heart rate, and sleep","🔴 Critical",14,[0,0,1,0,0,0],"📋 Planned"),
    ("Wearables","Garmin & Fitbit API integration","Expands compatibility for users with dedicated fitness trackers","🟠 High",10,[0,0,1,0,0,0],"📋 Planned"),
    ("Wearables","Continuous glucose monitor (CGM) feed","Automates blood glucose tracking for accurate metabolic simulation","🟠 High",8,[0,0,1,1,0,0],"📋 Planned"),
    ("Wearables","Wearable → BioGears real-time input","Uses live patient data to continuously recalibrate the twin's state","🔴 Critical",12,[0,0,1,1,0,0],"📋 Planned"),
    ("Clinical Data","HL7 FHIR record import","Pulls in verified clinical history from the user's healthcare provider","🟠 High",10,[0,0,1,0,0,0],"📋 Planned"),
    ("Clinical Data","PDF lab report OCR + parsing","Allows manual import of blood work and lab results from photos","🟢 Medium",8,[0,0,1,0,0,0],"📋 Planned"),
    ("Mobile App","Offline mode with local SQLite cache","Ensures app remains functional in areas with poor internet connection","🟠 High",7,[0,0,1,0,0,0],"📋 Planned"),

    # ── PHASE 4: AI & Personalisation (Aug) ─────────────────────────────────
    ("PHASE 4: AI & Personalisation","","","","","","","","","","","",""),
    ("AI / Analytics","Personalised risk score engine","Gives the user a simple, holistic metric of their overall health","🔴 Critical",14,[0,0,0,1,0,0],"📋 Planned"),
    ("AI / Analytics","Symptom → differential diagnosis RAG","Provides context-aware possible conditions when a user feels unwell","🔴 Critical",12,[0,0,0,1,0,0],"📋 Planned"),
    ("AI / Analytics","Medication interaction checker","Warns users of dangerous side effects when mixing prescriptions","🟠 High",10,[0,0,0,1,0,0],"📋 Planned"),
    ("AI / Analytics","Predictive health trend forecasting","Shows the user where their health is heading in the next 3-6 months","🟠 High",14,[0,0,0,1,1,0],"📋 Planned"),
    ("Mobile App","AI-generated weekly health report PDF","Summarizes weekly progress and simulation insights into a shareable format","🟠 High",8,[0,0,0,1,0,0],"📋 Planned"),
    ("Mobile App","Smart notification triggers (anomaly alert)","Actively alerts the user when their digital twin enters a dangerous state","🟢 Medium",6,[0,0,0,1,0,0],"📋 Planned"),
    ("Mobile App","Multilingual support (5 languages)","Broadens user base to non-English speaking demographics","🟢 Medium",10,[0,0,0,1,0,0],"📋 Planned"),

    # ── PHASE 5: Family & Social Features (Sep) ─────────────────────────────
    ("PHASE 5: Family & Social Features","","","","","","","","","","","",""),
    ("Family Health","Family dashboard (shared vitals view)","Allows monitoring the health of aging parents or dependents","🟠 High",10,[0,0,0,0,1,0],"📋 Planned"),
    ("Family Health","Guardian alerts for elderly/child members","Notifies a family member instantly if a loved one's vitals drop","🟠 High",8,[0,0,0,0,1,0],"📋 Planned"),
    ("Family Health","Role-based access control (RBAC)","Ensures privacy by controlling what data family or doctors can see","🟠 High",7,[0,0,0,0,1,0],"📋 Planned"),
    ("Family Health","Shared medication schedule","Helps caretakers ensure dependents are taking their medication on time","🟢 Medium",6,[0,0,0,0,1,0],"📋 Planned"),
    ("Clinical","Telemedicine / doctor share report flow","Easily sends the digital twin's data directly to a physician","🟢 Medium",10,[0,0,0,0,1,0],"📋 Planned"),
    ("Clinical","Emergency SOS with location + vitals","Saves lives by dispatching paramedics with the user's live health data","🔴 Critical",8,[0,0,0,0,1,0],"📋 Planned"),
    ("Security","End-to-end encryption (HIPAA prep)","Critical requirement for safely storing and transmitting medical data","🔴 Critical",12,[0,0,0,0,1,0],"📋 Planned"),

    # ── PHASE 6: Scale, Launch & Monetisation (Oct) ─────────────────────────
    ("PHASE 6: Scale, Launch & Monetisation","","","","","","","","","","","",""),
    ("Infrastructure","Kubernetes auto-scaling for BioGears","Handles high user load by spinning up simulation servers dynamically","🟠 High",14,[0,0,0,0,0,1],"📋 Planned"),
    ("Infrastructure","CI/CD pipeline (GitHub Actions + EAS)","Automates testing and deployment to speed up development cycles","🟠 High",8,[0,0,0,0,0,1],"📋 Planned"),
    ("Infrastructure","HIPAA compliance audit & penetration test","Legally required to store patient data and operate in the healthcare space","🔴 Critical",14,[0,0,0,0,0,1],"📋 Planned"),
    ("Infrastructure","Production monitoring (Grafana/Sentry)","Allows the engineering team to detect and fix bugs in real-time","🟠 High",8,[0,0,0,0,0,1],"📋 Planned"),
    ("Monetisation","Freemium tier + subscription paywall","Generates recurring revenue to sustain server and development costs","🟠 High",10,[0,0,0,0,1,1],"📋 Planned"),
    ("Monetisation","B2B clinic / hospital dashboard","Expands revenue by selling the platform to healthcare providers","🟢 Medium",14,[0,0,0,0,0,1],"📋 Planned"),
    ("Launch","App Store & Play Store submission","Makes the product officially available to the public for download","🔴 Critical",5,[0,0,0,0,0,1],"📋 Planned"),
    ("Launch","Beta user onboarding & feedback loop","Gathers critical user insights before the massive public launch","🟠 High",7,[0,0,0,0,0,1],"📋 Planned"),
]

month_bg = {0:"FFFFFF", 1:None}  # will use month colours when filled

row_idx = 4
feat_no = 0
for r in rows:
    cat, task, reason, prio, effort, *rest = r[0], r[1], r[2], r[3], r[4], r[5:]

    # Phase header rows
    if task == "":
        ws1.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
        c = ws1.cell(row_idx, 1, f"  {cat}")
        c.fill = fill("1A1A2E"); c.font = font("FFD700", True, 12)
        c.alignment = align("left"); c.border = border()
        ws1.row_dimensions[row_idx].height = 24
        row_idx += 1
        continue

    feat_no += 1
    gantt = rest[0][0]  # list of 6 ints
    status = rest[0][1]

    bg = C["alt_bg"] if feat_no % 2 == 0 else "FFFFFF"

    cell(ws1, row_idx, 1, feat_no, bg, h="center")
    cell(ws1, row_idx, 2, cat, bg)
    cell(ws1, row_idx, 3, task, bg, bold=True)
    cell(ws1, row_idx, 4, reason, bg)
    pb = prio_bg.get(prio, "FFFFFF")
    cell(ws1, row_idx, 5, prio, pb, h="center")
    cell(ws1, row_idx, 6, effort, bg, h="center")

    # Gantt bars
    for mi, filled in enumerate(gantt):
        mc = month_cols[mi].lstrip("#") if hasattr(month_cols[mi], "lstrip") else month_cols[mi]
        bar_bg = month_cols[mi] if filled else "F5F5F5"
        bar_txt = "██" if filled else ""
        bar_fg = "FFFFFF" if filled else "F5F5F5"
        cell(ws1, row_idx, 7+mi, bar_txt, bar_bg, bar_fg, h="center")

    # status
    st_bg = {"✅ Done": C["done"], "🔄 In Progress": C["wip"], "📋 Planned": C["todo"]}.get(status, "FFFFFF")
    cell(ws1, row_idx, 13, status, st_bg, h="center")

    ws1.row_dimensions[row_idx].height = 18
    row_idx += 1

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 – SPRINT PLAN
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("🗓 Sprint Plan")

ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "VitalHealth – 2-Week Sprint Breakdown (Phase 1 & 2 detailed)"
t2.fill = fill("0F3460"); t2.font = font("FFFFFF", True, 14)
t2.alignment = align("center"); ws2.row_dimensions[1].height = 30

sprint_cols = ["Sprint","Dates","Focus","Key Deliverables","Owner","Effort","Done?"]
sprint_widths = [8,20,25,50,15,10,10]
for i,(col,w) in enumerate(zip(sprint_cols,sprint_widths),1):
    hdr(ws2, 2, i, col, "1A1A2E")
    ws2.column_dimensions[get_column_letter(i)].width = w

sprints = [
    ("S-01","May 11–24","Engine Stability",
     "Fix pathing bugs, silent crash detection, XML validation, concurrency guard","Backend",18,"🔄"),
    ("S-02","May 25–Jun 7","Mobile Hardening",
     "Background job persistence, push notification on job complete, BioGears streaming","Full-Stack",16,"📋"),
    ("S-03","Jun 8–21","Substance & PK",
     "50+ drug registry, multi-substance dosing, PK unit validation, glucose model","Backend",20,"📋"),
    ("S-04","Jun 22–Jul 5","Real-Time Vitals",
     "SSE vitals stream, SSE mobile client, anomaly ML model v1","Full-Stack",18,"📋"),
    ("S-05","Jul 6–19","Wearables I",
     "Apple Health sync, Google Fit sync, Garmin API, wearable → BioGears input","Mobile",20,"📋"),
    ("S-06","Jul 20–Aug 2","Clinical Data",
     "HL7 FHIR import, PDF OCR lab reports, offline SQLite cache","Full-Stack",16,"📋"),
    ("S-07","Aug 3–16","AI Engine",
     "Personalised risk score, symptom RAG diagnosis, medication interaction checker","AI/ML",22,"📋"),
    ("S-08","Aug 17–30","Predictive Health",
     "Trend forecasting model, weekly PDF report gen, smart anomaly alerts","AI/ML",20,"📋"),
    ("S-09","Sep 1–14","Family Features",
     "Family dashboard, guardian alerts, RBAC, shared medication schedule","Full-Stack",18,"📋"),
    ("S-10","Sep 15–28","Compliance & SOS",
     "SOS with vitals, E2E encryption, HIPAA controls, telemedicine share flow","Security",20,"📋"),
    ("S-11","Oct 1–14","Scale & Infra",
     "Kubernetes autoscaling, CI/CD pipeline, Grafana monitoring, Sentry","DevOps",18,"📋"),
    ("S-12","Oct 15–31","Launch & Monetise",
     "Subscription paywall, App Store submission, B2B dashboard, beta feedback loop","All",22,"📋"),
]

for i, (sp, dates, focus, deliverables, owner, effort, done) in enumerate(sprints, 3):
    bg = "F0F8FF" if i % 2 == 0 else "FFFFFF"
    done_bg = {"✅": C["done"], "🔄": C["wip"], "📋": C["todo"]}.get(done, "FFFFFF")
    cell(ws2, i, 1, sp, bg, h="center")
    cell(ws2, i, 2, dates, bg, h="center")
    cell(ws2, i, 3, focus, bg, bold=True)
    cell(ws2, i, 4, deliverables, bg)
    cell(ws2, i, 5, owner, bg, h="center")
    cell(ws2, i, 6, f"{effort} days", bg, h="center")
    cell(ws2, i, 7, done, done_bg, h="center")
    ws2.row_dimensions[i].height = 22

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 – FEATURE PRIORITY MATRIX
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("⚡ Priority Matrix")

ws3.merge_cells("A1:F1")
t3 = ws3["A1"]
t3.value = "Feature Priority Matrix – Impact vs Effort"
t3.fill = fill("533483"); t3.font = font("FFFFFF", True, 14)
t3.alignment = align("center"); ws3.row_dimensions[1].height = 30

pm_cols = ["Feature","Category","Impact (1-10)","Effort (1-10)","Priority Score","Quadrant"]
pm_widths = [45,20,14,14,16,20]
for i,(col,w) in enumerate(zip(pm_cols,pm_widths),1):
    hdr(ws3, 2, i, col, "2C3E50")
    ws3.column_dimensions[get_column_letter(i)].width = w

matrix_rows = [
    ("Fix BioGears silent crash & pathing","Engine",10,3,None,"🚀 Quick Win"),
    ("Real-time vitals streaming (SSE)","Engine",9,5,None,"🚀 Quick Win"),
    ("Wearable → BioGears live input","Wearables",10,7,None,"💎 Strategic"),
    ("Personalised risk score engine","AI/ML",10,8,None,"💎 Strategic"),
    ("Symptom RAG differential diagnosis","AI/ML",9,7,None,"💎 Strategic"),
    ("HL7 FHIR clinical record import","Clinical",8,6,None,"💎 Strategic"),
    ("Apple Health / Google Fit sync","Wearables",9,5,None,"🚀 Quick Win"),
    ("Blood glucose & insulin model","Engine",8,7,None,"💎 Strategic"),
    ("Medication interaction checker","AI/ML",9,6,None,"💎 Strategic"),
    ("Emergency SOS with vitals","Clinical",10,4,None,"🚀 Quick Win"),
    ("HIPAA E2E encryption","Security",10,8,None,"💎 Strategic"),
    ("Family guardian alerts","Family",8,5,None,"🚀 Quick Win"),
    ("Weekly health report PDF","AI/ML",7,4,None,"🚀 Quick Win"),
    ("Subscription paywall","Monetisation",9,5,None,"🚀 Quick Win"),
    ("B2B clinic dashboard","Monetisation",8,8,None,"🎯 Long-term"),
    ("Multilingual support","UX",6,6,None,"🎯 Long-term"),
    ("Kubernetes autoscaling","Infra",8,7,None,"💎 Strategic"),
    ("CI/CD pipeline","Infra",7,4,None,"🚀 Quick Win"),
    ("PDF OCR lab report parsing","Clinical",7,5,None,"🚀 Quick Win"),
    ("Offline SQLite cache","Mobile",7,5,None,"🚀 Quick Win"),
]

quad_bg = {
    "🚀 Quick Win":"C8E6C9",
    "💎 Strategic":"BBDEFB",
    "🎯 Long-term":"FFF9C4",
    "⚠️ Re-evaluate":"FFCCBC",
}

for i, (feat, cat, impact, effort_val, _, quad) in enumerate(matrix_rows, 3):
    score = round((impact * 2 - effort_val) / 10 * 10, 1)
    bg = "F8F8FF" if i % 2 == 0 else "FFFFFF"
    qb = quad_bg.get(quad, "FFFFFF")
    cell(ws3, i, 1, feat, bg, bold=True)
    cell(ws3, i, 2, cat, bg, h="center")
    cell(ws3, i, 3, impact, bg, h="center")
    cell(ws3, i, 4, effort_val, bg, h="center")
    cell(ws3, i, 5, score, bg, h="center")
    cell(ws3, i, 6, quad, qb, h="center")
    ws3.row_dimensions[i].height = 18

# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 – KPI TARGETS
# ══════════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet("📊 KPI Targets")

ws4.merge_cells("A1:G1")
t4 = ws4["A1"]
t4.value = "VitalHealth – KPI & Success Metrics by Month"
t4.fill = fill("E94560"); t4.font = font("FFFFFF", True, 14)
t4.alignment = align("center"); ws4.row_dimensions[1].height = 30

kpi_cols = ["KPI / Metric","Baseline","May","Jun","Jul","Aug","Sep","Oct (Target)"]
kpi_widths = [38,14,10,10,10,10,10,14]
for i,(col,w) in enumerate(zip(kpi_cols,kpi_widths),1):
    hdr(ws4, 2, i, col, "1A1A2E")
    ws4.column_dimensions[get_column_letter(i)].width = w

kpis = [
    ("BioGears Sim Success Rate (%)","~55%","85%","92%","95%","97%","98%","99%"),
    ("Avg Sim Completion Time (min)","8–12 min","6 min","4 min","3 min","2 min","2 min","<2 min"),
    ("App Crash Rate (%)","~5%","3%","2%","1.5%","1%","0.5%","<0.5%"),
    ("Wearable Data Sources Supported","1","1","1","4","6","8","10+"),
    ("Substances Supported in Engine","~20","20","50","60","70","70","80+"),
    ("AI Diagnosis Accuracy (RAG %)","N/A","N/A","N/A","N/A","75%","82%","88%"),
    ("Active Beta Users","0","10","50","150","400","800","2,000"),
    ("Family Members Per Account (avg)","1","1","1.2","1.5","2","2.5","3+"),
    ("Push Notification Open Rate (%)","N/A","30%","35%","38%","40%","42%","45%"),
    ("Monthly Active Users (MAU)","0","20","80","200","500","1,200","3,000"),
    ("App Store Rating","N/A","N/A","N/A","N/A","N/A","N/A","4.6+"),
    ("Revenue (MRR $)","$0","$0","$0","$0","$500","$2,000","$8,000"),
]

for i, row in enumerate(kpis, 3):
    bg = "F0FFF4" if i % 2 == 0 else "FFFFFF"
    for j, val in enumerate(row, 1):
        bold = j == 1
        cell(ws4, i, j, val, bg, bold=bold, h="center" if j > 1 else "left")
    ws4.row_dimensions[i].height = 18

# ── Save ──────────────────────────────────────────────────────────────────────
out = "/home/akhilreddy/health-digital-twin/VitalHealth_6Month_Roadmap.xlsx"
wb.save(out)
print(f"✅  Saved → {out}")
